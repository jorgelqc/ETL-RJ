# Librerias usadas
import pandas as pd
from openpyxl import load_workbook
import datetime
import re
from sqlalchemy import create_engine, text
from sqlalchemy.exc import ProgrammingError, IntegrityError
import pyodbc
from tkinter import Tk, filedialog
import sys
import os
from dotenv import load_dotenv

def get_env_path():
    """Obtiene la ruta correcta del archivo .env."""
    if getattr(sys, 'frozen', False):
        # Entorno PyInstaller
        return os.path.join(sys._MEIPASS, '.env')
    else:
        # Entorno de desarrollo
        return '.env'

# Carga las variables de entorno
env_path = get_env_path()
load_dotenv(dotenv_path=env_path)

# --- Configuración de la Base de Datos ---
SERVER_NAME = os.environ.get("SERVER_NAME")
PORT = os.environ.get("PORT")
DATABASE_NAME = os.environ.get("DATABASE_NAME")
USERNAME = os.environ.get("DB_USERNAME")
PASSWORD = os.environ.get("DB_PASSWORD")
SERVER_AND_PORT = f"{SERVER_NAME}:{PORT}"

# --- Mapeos Estáticos ---
PRODUCTO_MAPPING = {
    'Ricky Joy Yogurt': 1,
    'Mellow Cones': 2,
    'Crazy Legs': 3,
    'Ricky Joy Gels': 4,
    'Jelly Fruits': 5,
    'Plis': 6,
    'SSC Roll On': 7,
    'Freeze Dried': 8,
    '3D Gummies': 9,
    'SC Gel': 10,
    'Cotton Candy': 11
}

ZONE_MAPPING = {
    'Zone 1': 1,
    'Zone 2': 2,
    'Zone 3': 3,
    'Zone 4': 4,
    'Zone 5': 5,
    'Zone 6': 6,
    'Zone 7': 7,
    'KamCentral': 8,
    'KamEast': 9,
    'E-Commerce': 10,
    'Outlet & Donation': 11
}

# --- Diccionarios de Meses y Año Actual ---
meses_en_a_es = {
    "January": "Enero", "February": "Febrero", "March": "Marzo",
    "April": "Abril", "May": "Mayo", "June": "Junio",
    "July": "Julio", "August": "Agosto", "September": "Septiembre",
    "October": "Octubre", "November": "Noviembre", "December": "Diciembre"
}
meses_a_numero = {
    'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4, 'Mayo': 5, 'Junio': 6,
    'Julio': 7, 'Agosto': 8, 'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
}
# Diccionario inverso para traducir de español a inglés
meses_es_a_en = {v: k for k, v in meses_en_a_es.items()}

año_actual = datetime.datetime.now().year

# --- Selección de Archivo ---
root = Tk()
root.withdraw()
print("Por favor, selecciona el archivo 'WOR Ventas.xlsx'...")
file_path = filedialog.askopenfilename(
    title="Selecciona el archivo 'WOR Ventas.xlsx'",
    filetypes=[("Archivos de Excel", "*.xlsx *.xls")]
)

if not file_path:
    print("No se seleccionó ningún archivo. Saliendo del programa.")
    exit()

print(f"Archivo seleccionado: {file_path}")
try:
    workbook = load_workbook(file_path, data_only=True)
    print("Archivo de Excel cargado exitosamente.")
except Exception as e:
    print(f"Error al cargar el archivo de Excel: {e}")
    exit()

# --- Patrones de Búsqueda Generalizados ---
todos_los_meses_pattern = "|".join(meses_a_numero.keys())

patrones = [
    re.compile(rf"Avancedeventa_Category_(?P<zona>Zone[1-6]|KamEast|KamCentral)_(?P<mes>{todos_los_meses_pattern})", re.IGNORECASE),
    re.compile(rf"Proyeccion_Vendedor_(?P<zona>Zone[1-6]|KamEast|KamCentral)_(?P<mes>{todos_los_meses_pattern})", re.IGNORECASE),
    re.compile(rf"Forecast_(?P<zona>Zone[1-6]|KamEast|KamCentral)_(?P<mes>{todos_los_meses_pattern})", re.IGNORECASE)
]

# Diccionario para almacenar todas las tablas encontradas
tablas_extraidas = {
    'category': {},
    'forecast': {},
    'zone_quotas': {}
}

# --- Bucle de Extracción Mejorado y CORREGIDO ---
print("\nBuscando y extrayendo tablas de todos los meses...")
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    if not sheet.tables:
        continue

    for table_name in sheet.tables:
        for patron in patrones:
            match = patron.fullmatch(table_name)
            if match:
                # Extraemos el mes en español que encontró el patrón
                nombre_mes_espanol = match.group('mes').capitalize()
                numero_mes_encontrado = meses_a_numero.get(nombre_mes_espanol)
                
                # Traducir el mes a inglés
                nombre_mes_ingles = meses_es_a_en.get(nombre_mes_espanol, nombre_mes_espanol)
                
                # ASUNCIÓN: Se asume que todas las tablas pertenecen al año actual.
                año_encontrado = año_actual
                
                # Obtener el objeto de la tabla usando su nombre
                table_object = sheet.tables[table_name]
                table_ref = table_object.ref
                
                data = sheet[table_ref]
                rows = [[cell.value for cell in row] for row in data]
                df = pd.DataFrame(rows[1:], columns=rows[0])

                # Añadir el nombre del mes en INGLÉS
                df['nombre_mes'] = nombre_mes_ingles
                df['mes'] = numero_mes_encontrado
                df['año'] = año_encontrado
                
                # Clasificar el DataFrame
                if 'Avancedeventa_Category' in table_name:
                    tablas_extraidas['category'][table_name] = df
                elif 'Forecast' in table_name:
                    tablas_extraidas['forecast'][table_name] = df
                
                print(f" -> Encontrada: {table_name}")
                # --- MENSAJE DE VERIFICACIÓN ---
                print(f"   -> Traduciendo mes: '{nombre_mes_espanol}' -> '{nombre_mes_ingles}'")
                break

# --- Funciones de Procesamiento y Limpieza ---
def procesar_cuotas_zona(df, nombre_tabla):
    """
    Procesa la primera fila de las tablas de forecast para extraer cuotas por zona
    """
    df_clean = df.copy()
    
    # Verificar que existe la columna TOTAL
    if 'TOTAL' not in df_clean.columns:
        print(f"Advertencia: No se encontró columna 'TOTAL' en {nombre_tabla}")
        return df_clean
    
    # Limpiar valores nulos en TOTAL
    df_clean['TOTAL'] = pd.to_numeric(df_clean['TOTAL'], errors='coerce').fillna(0)
    
    # Filtrar solo filas donde TOTAL > 0 (cuotas reales de zona)
    df_clean = df_clean[df_clean['TOTAL'] > 0]
    
    return df_clean

# --- MODIFICACIÓN 4: Procesar las cuotas de zona extraídas ---
print("\nProcesando cuotas de zona...")
for nombre_tabla, df in tablas_extraidas['zone_quotas'].items():
    df = procesar_cuotas_zona(df, nombre_tabla)
    tablas_extraidas['zone_quotas'][nombre_tabla] = df

total_zone_quotas = pd.concat(tablas_extraidas['zone_quotas'].values(), ignore_index=True) if tablas_extraidas['zone_quotas'] else pd.DataFrame()

if not total_zone_quotas.empty:
    total_zone_quotas = total_zone_quotas.rename(columns={"TOTAL": "cuota"}, errors='ignore')
    print(f"Se procesaron {len(total_zone_quotas)} cuotas de zona")

def ingest_zone_quotas_data(df_to_ingest):
    """
    Carga las cuotas generales por zona en la tabla Cuota_forecast
    """
    table_name = 'Cuota_forecast'
    
    if df_to_ingest.empty:
        print(f"\nDataFrame para cuotas de zona está vacío.")
        return
    
    engine = None
    try:
        connection_string = f"mssql+pymssql://{USERNAME}:{PASSWORD}@{SERVER_AND_PORT}/{DATABASE_NAME}"
        engine = create_engine(connection_string)
        print(f"\n--- Iniciando proceso de CUOTAS DE ZONA para '{table_name}' ---")
        
        df = df_to_ingest.copy()
        
        # Mapeo de Zonas (no necesita clientes para cuotas de zona)
        df['id_zone'] = df['Zone'].map(ZONE_MAPPING).fillna(1).astype(int)
        
        # Para cuotas de zona, el id_cliente será NULL o un valor especial (ej: 0)
        df['id_cliente'] = 0  # O puedes usar NULL si tu BD lo permite
        
        # Limpieza y preparación
        if 'cuota' not in df.columns and 'TOTAL' in df.columns:
            df = df.rename(columns={"TOTAL": "cuota"})
        
        df['cuota'] = pd.to_numeric(df['cuota'], errors='coerce').fillna(0).astype(float)
        
        # Filtrar solo cuotas válidas (mayor a 0)
        df = df[df['cuota'] > 0]
        
        cols_finales = ['id_zone', 'id_cliente', 'cuota', 'nombre_mes', 'mes', 'año']
        df = df.filter(items=cols_finales)
        
        # Lógica de Deduplicación específica para cuotas de zona
        unique_cols = ['id_zone', 'mes', 'año']
        query = f"SELECT {', '.join(unique_cols)} FROM {table_name} WHERE id_cliente = 0"
        existing_records_df = pd.read_sql_query(query, engine)
        
        if not existing_records_df.empty:
            merged = df.merge(existing_records_df, on=unique_cols, how='left', indicator=True)
            df_to_insert = merged[merged['_merge'] == 'left_only'].drop(columns=['_merge'])
        else:
            df_to_insert = df
            
        print(f"Total de cuotas de zona encontradas: {len(df)}")
        print(f"Cuotas de zona a insertar (nuevas): {len(df_to_insert)}")
        
        if not df_to_insert.empty:
            df_to_insert.to_sql(table_name, con=engine, if_exists='append', index=False)
            print(f"Se insertaron {len(df_to_insert)} cuotas de zona en '{table_name}'.")
        else:
            print("No hay cuotas de zona nuevas para insertar.")
            
    except Exception as e:
        print(f"\n¡ERROR en el proceso de cuotas de zona: {e}")
    finally:
        if engine: 
            engine.dispose()

def procesar_tabla(df, renombres_por_posicion):
    columnas = list(df.columns)
    for idx, nuevo_nombre in renombres_por_posicion.items():
        if idx < len(columnas):
            columnas[idx] = nuevo_nombre
    df.columns = columnas
    return df

def limpiar_dataframe(df, tipo_tabla):
    df_clean = df.copy().fillna(0)
    if tipo_tabla == 'forecast':
        df_clean = df_clean.drop(columns=['Py %'], errors='ignore')
        df_clean = df_clean.drop(index=df_clean.index[0], errors='ignore')
        df_clean = df_clean[~df_clean.apply(lambda row: row.astype(str).str.contains('Total').any(), axis=1)]
        columna_1 = df_clean.columns[0]
        df_clean = df_clean[df_clean[columna_1] != 0]
    return df_clean

def agregar_columna_zona(df, nombre_tabla):
    match = re.search(r'(Zone\s*\d+|KamEast|KamCentral)', nombre_tabla, re.IGNORECASE)
    if match:
        zona_encontrada = match.group(0).replace(" ", "") # Ej: "Zone1", "KamEast"
        # Normalizar a formato del ZONE_MAPPING
        if 'zone' in zona_encontrada.lower():
            df["Zone"] = f"Zone {zona_encontrada[-1]}"
        else:
            df["Zone"] = zona_encontrada
    return df

# --- Aplicar procesamiento a los DataFrames extraídos ---
nuevos_nombres_columnas = {3: "cuota_dinero", 4: "cuota_volumen"}

for nombre_tabla, df in tablas_extraidas['category'].items():
    df = procesar_tabla(df, nuevos_nombres_columnas)
    df = limpiar_dataframe(df, 'category')
    df = agregar_columna_zona(df, nombre_tabla)
    tablas_extraidas['category'][nombre_tabla] = df

for nombre_tabla, df in tablas_extraidas['forecast'].items():
    df = limpiar_dataframe(df, 'forecast')
    df = agregar_columna_zona(df, nombre_tabla)
    tablas_extraidas['forecast'][nombre_tabla] = df

# --- Apilar y Renombrar ---
total_Forecast = pd.concat(tablas_extraidas['forecast'].values(), ignore_index=True) if tablas_extraidas['forecast'] else pd.DataFrame()
total_category = pd.concat(tablas_extraidas['category'].values(), ignore_index=True) if tablas_extraidas['category'] else pd.DataFrame()

if not total_Forecast.empty:
    total_Forecast = total_Forecast.rename(columns={"ZONA/CLIENTE": "nombre_cliente", "WEEK 1": "semana_1", "WEEK 2": "semana_2", "WEEK 3": "semana_3", "WEEK 4": "semana_4", "WEEK 5": "semana_5"}, errors='ignore')
if not total_category.empty:
    total_category = total_category.rename(columns={"Negocio.": "nombre_producto", "Vta $": "cuota_dinero", "Vta Vol": "cuota_volumen"}, errors='ignore')

# --- FUNCIONES DE CARGA A BASE DE DATOS ---

def ingest_forecast_data(df_to_ingest):
    table_name = 'Forecast'
    if df_to_ingest.empty:
        print(f"\nDataFrame para la tabla '{table_name}' está vacío. No hay nada que insertar.")
        return

    engine = None
    try:
        connection_string = f"mssql+pymssql://{USERNAME}:{PASSWORD}@{SERVER_AND_PORT}/{DATABASE_NAME}"
        engine = create_engine(connection_string)
        print(f"\n--- Iniciando proceso para la tabla '{table_name}' ---")

        df = df_to_ingest.copy()
        
        # Mapeo de Clientes y Zonas
        with engine.connect() as connection:
            clientes_result = connection.execute(text("SELECT id_cliente, nombre_cliente FROM Clientes"))
            clientes_map = {str(row.nombre_cliente).strip().upper(): row.id_cliente for row in clientes_result}
        
        df['id_cliente'] = df['nombre_cliente'].str.strip().str.upper().map(clientes_map)
        df['id_zone'] = df['Zone'].map(ZONE_MAPPING).fillna(1).astype(int)
        df = df.dropna(subset=['id_cliente']).copy()
        df['id_cliente'] = df['id_cliente'].astype(int)

        # Limpieza y preparación
        cols_to_keep = ['semana_1', 'semana_2', 'semana_3', 'semana_4', 'semana_5', 'mes', 'año', 'id_cliente', 'id_zone', 'nombre_mes']
        df = df.filter(items=cols_to_keep)
        for col in ['semana_1', 'semana_2', 'semana_3', 'semana_4', 'semana_5']:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(float)
        
        # Lógica de Deduplicación
        unique_cols = ['id_cliente', 'id_zone', 'mes', 'año']
        existing_records_df = pd.read_sql_query(f"SELECT {', '.join(unique_cols)} FROM {table_name}", engine)
        
        if not existing_records_df.empty:
            merged = df.merge(existing_records_df, on=unique_cols, how='left', indicator=True)
            df_to_insert = merged[merged['_merge'] == 'left_only'].drop(columns=['_merge'])
        else:
            df_to_insert = df

        print(f"Total de filas encontradas: {len(df)}")
        print(f"Filas a insertar (nuevas): {len(df_to_insert)}")

        if not df_to_insert.empty:
            df_to_insert.to_sql(table_name, con=engine, if_exists='append', index=False, chunksize=1000)
            print(f"Se insertaron {len(df_to_insert)} registros en '{table_name}'.")

    except Exception as e:
        print(f"\n¡ERROR en el proceso para la tabla '{table_name}': {e}")
    finally:
        if engine: engine.dispose()

def ingest_cuotas_data(df_to_ingest):
    table_name = 'Cuotas_Avance_Categoria'
    if df_to_ingest.empty:
        print(f"\nDataFrame para la tabla '{table_name}' está vacío.")
        return

    engine = None
    try:
        connection_string = f"mssql+pymssql://{USERNAME}:{PASSWORD}@{SERVER_AND_PORT}/{DATABASE_NAME}"
        engine = create_engine(connection_string)
        print(f"\n--- Iniciando proceso para la tabla '{table_name}' ---")

        df = df_to_ingest.copy()
        
        # Mapeo de Productos y Zonas
        df['id_producto'] = df['nombre_producto'].str.strip().map(PRODUCTO_MAPPING)
        df['id_zone'] = df['Zone'].map(ZONE_MAPPING).fillna(1).astype(int)
        df = df.dropna(subset=['id_producto']).copy()
        df['id_producto'] = df['id_producto'].astype(int)
        
        # Limpieza y preparación
        cols_to_keep = ['cuota_dinero', 'cuota_volumen', 'id_producto', 'id_zone', 'nombre_mes', 'mes', 'año']
        df = df.filter(items=cols_to_keep)
        df['cuota_dinero'] = pd.to_numeric(df['cuota_dinero'], errors='coerce').fillna(0).astype(float)
        df['cuota_volumen'] = pd.to_numeric(df['cuota_volumen'], errors='coerce').fillna(0).astype(int)

        # Lógica de Deduplicación
        unique_cols = ['id_producto', 'id_zone', 'mes', 'año']
        existing_records_df = pd.read_sql_query(f"SELECT {', '.join(unique_cols)} FROM {table_name}", engine)
        
        if not existing_records_df.empty:
            merged = df.merge(existing_records_df, on=unique_cols, how='left', indicator=True)
            df_to_insert = merged[merged['_merge'] == 'left_only'].drop(columns=['_merge'])
        else:
            df_to_insert = df

        print(f"Total de filas encontradas: {len(df)}")
        print(f"Filas a insertar (nuevas): {len(df_to_insert)}")

        if not df_to_insert.empty:
            df_to_insert.to_sql(table_name, con=engine, if_exists='append', index=False, chunksize=1000)
            print(f"Se insertaron {len(df_to_insert)} registros en '{table_name}'.")

    except Exception as e:
        print(f"\n¡ERROR en el proceso para la tabla '{table_name}': {e}")
    finally:
        if engine: engine.dispose()

def ingest_cuota_forecast_data(df_to_ingest):
    table_name = 'Cuota_forecast'
    if df_to_ingest.empty or 'TOTAL' not in df_to_ingest.columns:
        print(f"\nDataFrame para '{table_name}' está vacío o no contiene la columna 'TOTAL'.")
        return

    engine = None
    try:
        connection_string = f"mssql+pymssql://{USERNAME}:{PASSWORD}@{SERVER_AND_PORT}/{DATABASE_NAME}"
        engine = create_engine(connection_string)
        print(f"\n--- Iniciando proceso para la tabla '{table_name}' ---")

        df = df_to_ingest.copy()

        # Mapeo de Clientes y Zonas
        with engine.connect() as connection:
            clientes_result = connection.execute(text("SELECT id_cliente, nombre_cliente FROM Clientes"))
            clientes_map = {str(row.nombre_cliente).strip().upper(): row.id_cliente for row in clientes_result}

        df['id_cliente'] = df['nombre_cliente'].str.strip().str.upper().map(clientes_map)
        df['id_zone'] = df['Zone'].map(ZONE_MAPPING).fillna(1).astype(int)
        df = df.dropna(subset=['id_cliente']).copy()
        df['id_cliente'] = df['id_cliente'].astype(int)
        
        # Limpieza y preparación
        df = df.rename(columns={"TOTAL": "cuota"})
        df['cuota'] = pd.to_numeric(df['cuota'], errors='coerce').fillna(0).astype(float)
        cols_finales = ['id_zone', 'id_cliente', 'cuota', 'nombre_mes', 'mes', 'año']
        df = df.filter(items=cols_finales)

        # Lógica de Deduplicación
        unique_cols = ['id_cliente', 'id_zone', 'mes', 'año']
        existing_records_df = pd.read_sql_query(f"SELECT {', '.join(unique_cols)} FROM {table_name}", engine)
        
        if not existing_records_df.empty:
            merged = df.merge(existing_records_df, on=unique_cols, how='left', indicator=True)
            df_to_insert = merged[merged['_merge'] == 'left_only'].drop(columns=['_merge'])
        else:
            df_to_insert = df

        print(f"Total de filas encontradas: {len(df)}")
        print(f"Filas a insertar (nuevas): {len(df_to_insert)}")

        if not df_to_insert.empty:
            df_to_insert.to_sql(table_name, con=engine, if_exists='append', index=False)
            print(f"Se insertaron {len(df_to_insert)} registros en '{table_name}'.")

    except Exception as e:
        print(f"\n¡ERROR en el proceso para la tabla '{table_name}': {e}")
    finally:
        if engine: engine.dispose()

# --- Ejecución del Proceso de Carga ---
print("\n" + "="*50)
print("INICIANDO PROCESO DE CARGA DE DATOS")
print("="*50)


ingest_zone_quotas_data(total_zone_quotas)
#ingest_cuota_forecast_data(total_Forecast)
ingest_forecast_data(total_Forecast)
ingest_cuotas_data(total_category)

print("\n" + "="*50)
print("PROCESO DE CARGA FINALIZADO")
print("="*50 + "\n")
